import os, time, urllib.request, openpyxl, operator, tweepy
from openpyxl import Workbook


class TwitterScrapper:
    def scrape(self, Consumer_Key, Consumer_Secret, Access_Token, Access_Token_Secret, tag, limit=20, lang='en'):
        print("Starting Scrapping Twitter")
        file_path = "data/data_" + tag
        keyword = tag

        consumerKey = Consumer_Key
        consumerSecret = Consumer_Secret
        accessToken = Access_Token
        accessTokenSecret = Access_Token_Secret

        auth = tweepy.OAuthHandler(consumerKey, consumerSecret)
        auth.set_access_token(accessToken, accessTokenSecret)
        api = tweepy.API(auth)

        tweets = tweepy.Cursor(api.search, q=tag, lang=lang).items(limit)

        # Create a workbook for excel
        tag_File = file_path + "/" + tag + "_Twitter.xlsx"
        wb = openpyxl.Workbook()
        ws_Captions = wb.create_sheet(title="Posts")
        col = 'A'
        row = 1

        img_src = []
        hashtags = {}
        ext_links = []
        for tweet in tweets:
            text = tweet.text.lower()
            ws_Captions[col + str(row)] = text

            # stripping for urls and hashtags and their frequencies
            for tag in text.split():
                if tag.startswith("#"):
                    if tag[1:] not in hashtags:
                        hashtags[tag[1:]] = 1
                    elif tag[1:] in hashtags:
                        hashtags[tag[1:]] = hashtags[tag[1:]] + 1
                if tag[:4] == 'http':
                    ext_links.append(tag)

            try:
                mu = tweet.entities['media'][0]['media_url']
                if (len(mu) > 1):
                    img_src.append(mu)
            except:
                pass

            if (row % 50 == 0):
                print("Dumped " + str(row) + " Tweets")
            row += 1


        hashtags = sorted(hashtags.items(), key=operator.itemgetter(1), reverse=True)

        ws_Tags = wb.create_sheet(title="Tags")
        tagName = 'A'
        tagFreq = 'B'
        row = 1

        print("Dumping Related Hashtags")
        for tag in hashtags:
            ws_Tags[tagName + str(row)] = tag[0]
            ws_Tags[tagFreq + str(row)] = tag[1]
            row += 1


        print("Dumping External Links")
        ws_Links = wb.create_sheet(title="Links")
        row = 1
        for link in ext_links:
            ws_Links['A' + str(row)] = link
            row += 1


        wb.save(tag_File)

        time.sleep(5)

        print("Dumping " + str(len(img_src)) + " Images")
        row = 1
        for src in img_src:
            try:
                print("(" + str(row) + "/" + str(len(img_src)) + ") Images Downloaded")
                urllib.request.urlretrieve(src, file_path + '/img/Twitter_' + str(row) + ".jpeg")
                row += 1
                time.sleep(1.5)
            except:
                print("Image Download Failed. Downloading next image")

        print("Closing Twitter")
        return hashtags